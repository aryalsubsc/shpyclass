import openpyxl as xl
from openpyxl.chart import BarChart,Reference
wb=xl.load_workbook('tranasaction.xlsx')
sheet=wb['Sheet1']

for row in range(1,sheet.max_row+1):
    cell=sheet.cell(row,1)
    correct_value=cell.value
    correct_cell=sheet.cell(row,3)
    correct_cell.value=correct_value
values=Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=2,max_col=2)
chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'d2')
wb.save('transaction2.xlsx')